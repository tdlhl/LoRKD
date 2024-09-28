import json

import pandas as pd
import openpyxl


def contains(text, key):
    if isinstance(key, str):
        return key in text
    elif isinstance(key, list):
        for k in key:
            if k in text:
                return True
        return False         
    
def merge_modality(mod):
    if contains(mod, ['t1', 't2', 'mri', 'flair', 'dwi']):
        return 'mri'
    if contains(mod, 'ct'):
        return 'ct'
    if contains(mod, 'pet'):
        return 'pet'
    if contains(mod, 'fundus'):
        return 'fundus'
    else:
        return mod
    
def merge_label(label):
    if label == 'left atrium':
        return 'left heart atrium'
    if label == 'adrenal':
        return 'adrenal gland'
    if label == 'hepatic tumor':
        return 'liver tumour'
    else:
        return label

def merge(mod_label_json, xlsx2load, xlsx2save):
    dataset2mod = {}
    mod_lab2dice = {}
    
    # Load the first sheet of the Excel file
    excel_file_path = xlsx2load 
    df = pd.read_excel(excel_file_path, sheet_name=0)
    has_nsd = True if len(df.columns) > 2 else False
    
    # 将Dataset Merged 写入新的工作表
    workbook = openpyxl.load_workbook(xlsx2save)
    new_sheet = workbook.create_sheet(title='Dataset Merge', index=1)
    new_sheet.cell(row=1, column=1, value='Dataset')
    new_sheet.cell(row=1, column=2, value='Dice')
    new_sheet.cell(row=1, column=3, value='NSD')
    row = 2
    for i in range(1, len(df)):
        if ',' not in df.iloc[i, 0]:
            new_sheet.cell(row=row, column=1, value=df.iloc[i, 0])
            new_sheet.cell(row=row, column=2, value=df.iloc[i, 1])
            if has_nsd:
                new_sheet.cell(row=row, column=3, value=df.iloc[i, 2])
            row += 1
    
    # with pd.ExcelWriter(xlsx2save, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    #     filtered_df.to_excel(writer, sheet_name='Dataset Merge', index=False)
    
    # 选取前两列
    dataset_label_ls = df.iloc[:, 0]
    dice_ls = df.iloc[:, 1]
    nsd_ls = df.iloc[:, 2] if has_nsd else [0] * len(df)
    
    for dataset_label, dice, nsd in zip(dataset_label_ls, dice_ls, nsd_ls):  # MSD_Pancreas, pancreas   0.89
        if ', ' not in dataset_label:
            continue
        dataset, label = dataset_label.split(', ')
        label = label.lower()   # pancreas
        label = merge_label(label)
        
        # retrieval modality
        if dataset not in dataset2mod:
            with open(f'/mnt/petrelfs/lihaolin/project/SAT-decompose/LoRKD-seg/dataset/test/subset/{dataset}.jsonl', 'r') as f:
                lines = f.readlines()
            datum = json.loads(lines[0])
            modality = datum['modality'].lower()
            dataset2mod[dataset] = merge_modality(modality) # ct
            
        # unique id : modality_label
        mod_lab = f'{dataset2mod[dataset]}_{label}'
        
        # accumulate : dice and where the dice comes from (dataset, label, modality)
        if mod_lab not in mod_lab2dice:
            mod_lab2dice[mod_lab] = {'dice':[], 'nsd':[], 'merge':[]}
        mod_lab2dice[mod_lab]['dice'].append(dice)
        mod_lab2dice[mod_lab]['nsd'].append(nsd)
        mod_lab2dice[mod_lab]['merge'].append(f'{dataset_label}({modality})')
        
    # retrieval regions
    with open(mod_label_json, 'r') as f:
        dict = json.load(f)
    region2label = dict['region_based']
    for region, label_ls in region2label.items():
        region2label[region] = [mod_lab.split('_')[-1] for mod_lab in label_ls] # 去除modality
    region2label['abnormal'] = [mod_lab.split('_')[-1] for mod_lab in dict['abnormal']]
        
    region_dice_ls = {k:[] for k in region2label.keys()} # {'brain':[0.9, ...], ...}
    region_nsd_ls = {k:[] for k in region2label.keys()} # {'brain':[0.9, ...], ...}
    region_merge_ls = {k:[] for k in region2label.keys()} # {'brain':['frontal lobe', ...], ...}
    
    mod_lab_ls = []
    dice_ls = []
    nsd_ls = []
    merge_ls = []  
    region_ls = []
    for mod_lab, dict in mod_lab2dice.items():
        label = mod_lab.split('_')[-1]
        mod_lab_ls.append(mod_lab)
        dice_ls.append(sum(dict['dice'])/len(dict['dice']))
        nsd_ls.append(sum(dict['nsd'])/len(dict['nsd']))
        merge_ls.append(','.join(dict['merge']))
        
        # find region
        if label in region2label['abnormal']:
            region_dice_ls['abnormal'].append(dice_ls[-1])
            region_nsd_ls['abnormal'].append(nsd_ls[-1])
            region_merge_ls['abnormal'].append(mod_lab)
            region_ls.append('abnormal')
        else:
            found = False
            for region, labels_in_region in region2label.items():
                if label in labels_in_region:
                    region_dice_ls[region].append(dice_ls[-1])
                    region_nsd_ls[region].append(nsd_ls[-1])
                    region_merge_ls[region].append(mod_lab)
                    region_ls.append(region)
                    found = True
                    break
            if not found:
                print(label)
                region_ls.append('unknown')
            
    df = pd.DataFrame({
        'Modality_Label': mod_lab_ls,
        'Dice': dice_ls,
        'NSD': nsd_ls,
        'Merge': merge_ls,
        'Region': region_ls
    })

    #book = openpyxl.load_workbook(xlsx2save)
    #writer = pd.ExcelWriter(xlsx2save, engine='openpyxl')
    #writer.book = book
    
    # with pd.ExcelWriter(xlsx2save, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    #     df.to_excel(writer, sheet_name='Label Merge', index=False)

    # 将Label Merged DataFrame写入新的工作表
    new_sheet = workbook.create_sheet(title='Label Merge', index=1)
    new_sheet.cell(row=1, column=1, value='Modality_Label')
    new_sheet.cell(row=1, column=2, value='Dice')
    new_sheet.cell(row=1, column=3, value='NSD')
    new_sheet.cell(row=1, column=4, value='Merge')
    new_sheet.cell(row=1, column=5, value='Region')
    row = 2
    for mod_lab, dice, nsd, merge, region in zip(mod_lab_ls, dice_ls, nsd_ls, merge_ls, region_ls):
        new_sheet.cell(row=row, column=1, value=mod_lab)
        new_sheet.cell(row=row, column=2, value=dice)
        new_sheet.cell(row=row, column=3, value=nsd)
        new_sheet.cell(row=row, column=4, value=merge)
        new_sheet.cell(row=row, column=4, value=region)
        row += 1
    
    # 将Region Merged 写入新的工作表
    new_sheet = workbook.create_sheet(title='Region Merge', index=1)
    new_sheet.cell(row=1, column=1, value='Region')
    new_sheet.cell(row=1, column=2, value='Dice')
    new_sheet.cell(row=1, column=3, value='NSD')
    new_sheet.cell(row=1, column=4, value='Merge')
    row = 2
    for key in region_dice_ls.keys():
        if len(region_dice_ls[key]) == 0:
            dice = nsd = 0
            merge = None
        else:
            dice = sum(region_dice_ls[key])/len(region_dice_ls[key])
            nsd = sum(region_nsd_ls[key])/len(region_nsd_ls[key])
            merge = ','.join(region_merge_ls[key])
        class_name = f'{key}({len(region_dice_ls[key])})'
        new_sheet.cell(row=row, column=1, value=class_name)
        new_sheet.cell(row=row, column=2, value=dice)
        new_sheet.cell(row=row, column=3, value=nsd)
        new_sheet.cell(row=row, column=4, value=merge)
        row += 1

    workbook.save(xlsx2save)
    
if __name__ == '__main__':
    import argparse
    
    def str2bool(v):
        if isinstance(v, bool):
            return v
        if v.lower() in ('yes', 'true', 't', 'y', '1'):
            return True
        elif v.lower() in ('no', 'false', 'f', 'n', '0'):
            return False
        else:
            raise argparse.ArgumentTypeError('Boolean value expected.')
    
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx2load', type=str,default='/mnt/petrelfs/lihaolin/project/SAT-decompose/LoRKD-seg/evaluate/class_results_v2_macro.csv')
    parser.add_argument('--xlsx2save', type=str, default='/mnt/petrelfs/lihaolin/project/SAT-decompose/LoRKD-seg/evaluate/class_results_nnunet.xlsx')
    parser.add_argument('--mod_lab_json', type=str, default='/mnt/petrelfs/lihaolin/project/SAT-decompose/LoRKD-seg/dataset/mod_lab(49).json')
    config = parser.parse_args()
    
    if not config.xlsx2save:
        config.xlsx2save = config.xlsx2load
    
    # region_merged(config.mod_lab_json, config.xlsx2load, config.xlsx2save)
    
    merge(config.mod_lab_json, config.xlsx2load, config.xlsx2save)            